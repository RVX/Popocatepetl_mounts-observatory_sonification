"""POPO01 -- Popocatepetl seismic/infrasound + satellite sonification toolkit.

Sonifies Popocatepetl (Mexico) volcanic activity from two independent data
sources: station MX.CZB ground seismic/infrasound data (provided by Sebastien
Valade, MOUNTS observatory) and the MOUNTS satellite API/export.

Two independent data sources, two independent sonification approaches:

  A. GROUND (seismic + infrasound), actions: fetch / plot / sonify / play / map
     Reads a local SDS (SeisComP Data Structure) archive -- there is no live
     FDSN server for this station yet. Sebastien sent one day of data
     (2026-03-27, the day of the last explosive activity, ~07:00 UTC) via
     WeTransfer; an FDSN web service may be set up later. Sonification method
     is "audification": keep every sample, just declare a faster playback
     sample rate so inaudible ground motion / pressure shifts up into the
     audible range.

  B. SATELLITE (MOUNTS API), action: satellite
     SO2 mass (Sentinel-5P), thermal hot-pixel count (Sentinel-2/MIROVA),
     InSAR deformation std.dev and coherence (Sentinel-1) -- four sparse,
     irregularly-sampled time series fetched over HTTP as JSON (or read from
     a local Excel export). There is no continuous waveform to "speed up"
     here, so this uses a different, complementary method: parameter-mapping
     sonification. Each measurement becomes a short pitched tone (pitch <-
     value, position in time <- measurement time, compressed into
     --listen-minutes), the classic "data sonification" approach for
     irregular scientific time series.

Station / data notes (from Sebastien Valade, March 2026)
----------------------------------------------------------
  Network = MX, Station = CZB
  Seismic channels  : HNZ, HNN, HNE  (location 00)   -- strong-motion/accelerometer
  Infrasound channel: HDF            (locations 01-04, four co-located sensors)
  "IP" (infrasonic parameter) on the MOUNTS website rises during ongoing
  eruptive activity -- that's when the infrasound channels are most
  interesting to listen to.
  Website: http://mounts-observatory.org/views/341090
  Popocatepetl MOUNTS/Smithsonian volcano id: 341090

Folder layout (created automatically next to this script)
-----------------------------------------------------------
  seed/                              unzip the WeTransfer SDS archive here
  datasets/ground/mseed/             processed ground waveform data (.mseed) + .json sidecar
  datasets/ground/plot/              ground spectrogram + waveform plots (.png)
  datasets/ground/sonifications/     ground audio (.wav)
  datasets/maps/                     station location plot (.png)
  datasets/satellite/raw/            raw MOUNTS API responses (.json)
  datasets/satellite/plot/           satellite time-series plots (.png)
  datasets/satellite/sonifications/  satellite parameter-mapping audio (.wav)
  datasets/satellite/envelopes/      per-series 0..1 control-curve CSVs, same
                                      timeline as the .wav -- for a REAPER
                                      envelope or a GPIO LED brightness script
                                      (see reaper/ and gpio/)

Examples
--------
  py POPO01.py                                          # ground: fetch+plot+sonify the
                                                          # eruption window (default)
  py POPO01.py --channels seismic fetch plot sonify      # seismic only (HNZ/HNN/HNE)
  py POPO01.py --channels infrasound fetch plot sonify   # infrasound only (HDF 01-04)
  py POPO01.py --full-day fetch plot                     # the entire archived day
  py POPO01.py --pick 0 sonify --channel HDF --speed-up 50
  py POPO01.py map                                       # station/volcano location plot
  py POPO01.py satellite                                 # fetch+plot+sonify all 4 MOUNTS series
  py POPO01.py satellite --sat-types so2,mirova --listen-minutes 3
"""
import os
import sys
import glob
import csv
import json
import math
import colorsys
import shutil
import argparse
import subprocess
import urllib.request
import urllib.parse
import email.utils
from datetime import datetime, timedelta, timezone

import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from scipy.signal import spectrogram
from scipy.io import wavfile

# obspy is only needed for the ground (seismic/infrasound) commands -- kept
# optional so `satellite` can run on lighter deployments (e.g. the
# Raspberry Pi's hourly refresh job) without installing it.
try:
    from obspy import UTCDateTime, read
    from obspy.clients.filesystem.sds import Client as SDSClient
except ImportError:
    UTCDateTime = read = SDSClient = None


def _require_obspy():
    if UTCDateTime is None:
        raise SystemExit(
            "obspy is required for ground-data commands (pip install obspy). "
            "Not needed for the `satellite` command."
        )

# ---------------------------------------------------------------------------
# Station / network configuration
# ---------------------------------------------------------------------------
NETWORK = "MX"
STATION = "CZB"

# Popocatepetl's Smithsonian/MOUNTS volcano id -- used both to label the ground
# plots and to query the satellite API (section B).
MOUNTS_TARGET_ID = 341090
VOLCANO_NAME = "Popocatepetl"
# Summit coordinates (public/Smithsonian GVP data). CZB's own exact
# coordinates were not included in the data package -- ask Sebastien Valade
# for the station StationXML/dataless if precise lat/lon/elevation/distance
# to the crater are needed for a future map revision.
VOLCANO_LAT = 19.023
VOLCANO_LON = -98.622
VOLCANO_ELEV_M = 5426

# Channel tokens the CLI understands, mapped to (location, channel) SEED codes.
# HN? = high broadband accelerometer (strong motion); HDF = infrasound (pressure).
CHANNEL_TOKEN_MAP = {
    "HNZ": ("00", "HNZ"),
    "HNN": ("00", "HNN"),
    "HNE": ("00", "HNE"),
    "HDF01": ("01", "HDF"),
    "HDF02": ("02", "HDF"),
    "HDF03": ("03", "HDF"),
    "HDF04": ("04", "HDF"),
}
SEISMIC_TOKENS = ["HNZ", "HNN", "HNE"]
INFRASOUND_TOKENS = ["HDF01", "HDF02", "HDF03", "HDF04"]

# Units reported after instrument-response removal, per channel family. If no
# inventory is supplied (--inventory), data stays in raw digitizer counts.
UNITS_BY_CHANNEL_PREFIX = {"HN": "m/s^2", "HDF": "Pa"}

# The data package covers exactly this one day, and this is the moment
# Sebastien flagged as the last explosive activity -- used as the default
# "interesting" window (see --around-eruption / --before-min / --after-min).
DEFAULT_DATE = "2026-03-27"
DEFAULT_ERUPTION_UTC_HHMM = "07:00"
DEFAULT_BEFORE_MIN = 30.0
DEFAULT_AFTER_MIN = 90.0

COMPONENT_COLORS = {
    "Z": "#00e5ff",   # cyan
    "N": "#ff9100",   # orange
    "E": "#ff1744",   # red/pink
    "F": "#b388ff",   # violet -- infrasound (HDF)
}

DEFAULT_SPEED_UP = 200
MIN_LISTEN_SPEED_UP = 100.0

BG_COLOR = "#0b0c10"
FG_COLOR = "0.85"
GRID_COLOR = "0.3"

# ---------------------------------------------------------------------------
# Folder layout
# ---------------------------------------------------------------------------
# Ground and satellite outputs are nested under their own subfolder
# (datasets/ground/*, datasets/satellite/*) so e.g. "sonifications" or "plot"
# never exists twice as an ambiguous top-level folder name.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_SDS_ROOT = os.path.join(BASE_DIR, "seed")
DATASETS_DIR = os.path.join(BASE_DIR, "datasets")
GROUND_DIR = os.path.join(DATASETS_DIR, "ground")
PLOT_DIR = os.path.join(GROUND_DIR, "plot")
MSEED_DIR = os.path.join(GROUND_DIR, "mseed")
SONIFY_DIR = os.path.join(GROUND_DIR, "sonifications")
MAP_DIR = os.path.join(DATASETS_DIR, "maps")
SAT_DIR = os.path.join(DATASETS_DIR, "satellite")
SAT_RAW_DIR = os.path.join(SAT_DIR, "raw")
SAT_PLOT_DIR = os.path.join(SAT_DIR, "plot")
SAT_SONIFY_DIR = os.path.join(SAT_DIR, "sonifications")
SAT_ENVELOPE_DIR = os.path.join(SAT_DIR, "envelopes")

for _dir in (PLOT_DIR, MSEED_DIR, SONIFY_DIR, MAP_DIR, SAT_RAW_DIR, SAT_PLOT_DIR, SAT_SONIFY_DIR, SAT_ENVELOPE_DIR):
    os.makedirs(_dir, exist_ok=True)



# ---------------------------------------------------------------------------
# Local civil time -- Mexico City is UTC-6 (CST) / UTC-5 (CDT), DST from
# first Sunday of April to last Sunday of October (Mexico's historical rule;
# implemented manually, no zoneinfo/tzdata dependency needed).
# ---------------------------------------------------------------------------
def _first_sunday(year, month):
    d = datetime(year, month, 1)
    while d.weekday() != 6:
        d += timedelta(days=1)
    return d


def _last_sunday(year, month):
    if month == 12:
        next_month_first = datetime(year + 1, 1, 1)
    else:
        next_month_first = datetime(year, month + 1, 1)
    d = next_month_first - timedelta(days=1)
    while d.weekday() != 6:
        d -= timedelta(days=1)
    return d


def _mx_offset_hours(utc_dt):
    year = utc_dt.year
    dst_start = _first_sunday(year, 4).replace(hour=2)
    dst_end = _last_sunday(year, 10).replace(hour=2)
    return -5 if dst_start <= utc_dt < dst_end else -6


def format_local_time(utc_dt):
    offset_h = _mx_offset_hours(utc_dt)
    local_dt = utc_dt + timedelta(hours=offset_h)
    tz_label = "CDT" if offset_h == -5 else "CST"
    return local_dt.strftime("%Y-%m-%d %H:%M"), tz_label


# ---------------------------------------------------------------------------
# Channel selection helpers
# ---------------------------------------------------------------------------
def parse_channels_arg(value):
    """Parse --channels ('seismic' | 'infrasound' | 'all' | comma-list of
    tokens from CHANNEL_TOKEN_MAP) into a list of tokens."""
    if value is None:
        value = "all"
    value = value.strip().lower()
    if value == "seismic":
        return list(SEISMIC_TOKENS)
    if value == "infrasound":
        return list(INFRASOUND_TOKENS)
    if value == "all":
        return list(SEISMIC_TOKENS) + list(INFRASOUND_TOKENS)
    tokens = [t.strip().upper() for t in value.split(",") if t.strip()]
    unknown = [t for t in tokens if t not in CHANNEL_TOKEN_MAP]
    if unknown:
        raise SystemExit(
            f"--channels: unknown token(s) {unknown}. Known: "
            f"{sorted(CHANNEL_TOKEN_MAP)}, or 'seismic'/'infrasound'/'all'."
        )
    return tokens


def units_for_trace(tr, response_removed):
    if not response_removed:
        return "counts"
    chan = tr.stats.channel.upper()
    for prefix, unit in UNITS_BY_CHANNEL_PREFIX.items():
        if chan.startswith(prefix):
            return unit
    return "counts"


# ---------------------------------------------------------------------------
# Time window helpers
# ---------------------------------------------------------------------------
def resolve_window(args):
    """Turn --date/--start/--end/--around-eruption/--full-day into a concrete
    (starttime, endtime) UTCDateTime pair."""
    _require_obspy()
    date_str = args.date or DEFAULT_DATE
    try:
        year, month, day = (int(p) for p in date_str.split("-"))
    except ValueError:
        raise SystemExit(f"--date must be YYYY-MM-DD (got {date_str!r})")
    day_start = UTCDateTime(year, month, day)

    if args.full_day:
        return day_start, day_start + 24 * 3600

    if args.start or args.end:
        if not (args.start and args.end):
            raise SystemExit("--start and --end must be given together (HH:MM, UTC)")
        starttime = _parse_hhmm(day_start, args.start)
        endtime = _parse_hhmm(day_start, args.end)
        if endtime <= starttime:
            raise SystemExit("--end must be after --start")
        return starttime, endtime

    # Default: a window around the known eruptive activity for this day.
    eruption = _parse_hhmm(day_start, DEFAULT_ERUPTION_UTC_HHMM)
    starttime = eruption - args.before_min * 60
    endtime = eruption + args.after_min * 60
    print(
        f"[info] No --start/--end/--full-day given; using the default window around "
        f"the {DEFAULT_ERUPTION_UTC_HHMM} UTC explosive activity on {date_str}: "
        f"{args.before_min:.0f} min before to {args.after_min:.0f} min after "
        f"(--around-eruption). Pass --start/--end or --full-day to override."
    )
    return starttime, endtime


def _parse_hhmm(day_start, hhmm):
    try:
        hh, mm = (int(p) for p in hhmm.split(":"))
    except ValueError:
        raise SystemExit(f"Expected HH:MM (UTC), got {hhmm!r}")
    return day_start + hh * 3600 + mm * 60


# ---------------------------------------------------------------------------
# Fetch (from a local SDS archive -- no live FDSN server for this station)
# ---------------------------------------------------------------------------
def build_sds_client(sds_root):
    _require_obspy()
    if not os.path.isdir(sds_root):
        raise SystemExit(
            f"SDS root not found: {sds_root}\n"
            f"Unzip Sebastien Valade's WeTransfer archive so it creates a "
            f"'{os.path.basename(sds_root)}/2026/MX/CZB/...' folder here, "
            f"or pass --sds-root PATH to point at wherever you unzipped it."
        )
    return SDSClient(sds_root=sds_root)


def fetch_window(client, channel_tokens, starttime, endtime):
    """Pull one Stream covering every requested channel token from the local
    SDS archive. Missing channels are warned about, not fatal -- e.g. asking
    for --full-day when the archive only has the eruption hours."""
    from obspy import Stream
    st = Stream()
    missing = []
    for token in channel_tokens:
        location, channel = CHANNEL_TOKEN_MAP[token]
        try:
            got = client.get_waveforms(
                network=NETWORK, station=STATION, location=location,
                channel=channel, starttime=starttime, endtime=endtime,
            )
        except Exception as exc:
            missing.append((token, str(exc)))
            continue
        if len(got) == 0:
            missing.append((token, "no data in archive for this window"))
            continue
        st += got
    if len(st) == 0:
        raise ValueError(
            f"No data found in SDS archive '{client.sds_root}' for "
            f"{starttime} - {endtime} (tried: {channel_tokens})"
        )
    if missing:
        print(f"[fetch] {len(missing)} requested channel(s) had no data: " +
              ", ".join(f"{t} ({e})" for t, e in missing))
    return st


def describe_stream(st, prefix="info"):
    print(f"[{prefix}] {len(st)} trace(s):")
    for tr in st:
        duration_s = tr.stats.npts / tr.stats.sampling_rate
        print(f"  - {tr.id}  |  {tr.stats.sampling_rate:.1f} Hz  |  "
              f"{duration_s:.1f}s  |  {tr.stats.npts} samples")


def load_stream(mseed_path):
    _require_obspy()
    if not os.path.exists(mseed_path):
        raise SystemExit(f"File not found: {mseed_path}")
    try:
        st = read(mseed_path)
    except Exception as exc:
        raise SystemExit(f"Could not read {mseed_path} as MSEED data: {exc}") from exc
    empty_ids = [tr.id for tr in st if tr.stats.npts == 0]
    if empty_ids:
        print(f"[warn] Dropping {len(empty_ids)} zero-length trace(s): {', '.join(empty_ids)}",
              file=sys.stderr)
        st.traces = [tr for tr in st if tr.stats.npts > 0]
    if len(st) == 0:
        raise SystemExit(f"{mseed_path} contains no usable (non-empty) traces.")
    return st


def _process_and_save(st, args, file_prefix, response_removed):
    """Detrend, optionally remove instrument response, bandpass per channel
    family (seismic vs. infrasound get different corners since infrasound
    signals of interest are generally lower-frequency), taper, save .mseed."""
    st.detrend("demean")
    st.detrend("linear")
    for tr in st.traces:
        is_infra = tr.stats.channel.upper().startswith("HDF")
        freqmin = args.infra_freqmin if is_infra else args.freqmin
        freqmax = args.infra_freqmax if is_infra else args.freqmax
        tr.filter("bandpass", freqmin=freqmin, freqmax=freqmax, corners=4)
    st.taper(0.125, max_length=20.0)

    starttime = min(tr.stats.starttime for tr in st)
    timestamp = starttime.strftime("%Y-%m-%d-%H-%M")
    mseed_path = os.path.join(MSEED_DIR, f"{file_prefix}_{timestamp}.mseed")
    st.write(mseed_path, format="MSEED")
    print(f"[fetch] Saved waveform data to {mseed_path}")
    describe_stream(st, prefix="fetch")

    metadata = dict(
        network=NETWORK, station=STATION, channels=[tr.id for tr in st],
        freqmin=args.freqmin, freqmax=args.freqmax,
        infra_freqmin=args.infra_freqmin, infra_freqmax=args.infra_freqmax,
        response_removed=response_removed,
        starttime=str(st[0].stats.starttime), endtime=str(st[0].stats.endtime),
    )
    metadata_path = os.path.splitext(mseed_path)[0] + ".json"
    with open(metadata_path, "w", encoding="utf-8") as f:
        json.dump(metadata, f, indent=2)
    return mseed_path


def do_fetch(args):
    channel_tokens = parse_channels_arg(args.channels)
    starttime, endtime = resolve_window(args)
    client = build_sds_client(args.sds_root)
    print(f"[fetch] Reading {NETWORK}.{STATION} from '{args.sds_root}': "
          f"{starttime} - {endtime} UTC, channels {channel_tokens}")
    st = fetch_window(client, channel_tokens, starttime, endtime)

    response_removed = False
    if args.inventory:
        from obspy import read_inventory
        if not os.path.exists(args.inventory):
            raise SystemExit(f"--inventory not found: {args.inventory}")
        inv = read_inventory(args.inventory)
        kept = []
        for tr in st.traces:
            try:
                tr.remove_response(inventory=inv)
                kept.append(tr)
            except Exception as exc:
                print(f"[warn] Could not remove response for {tr.id} ({exc}); "
                      f"keeping raw counts for this trace.", file=sys.stderr)
                kept.append(tr)
        st.traces = kept
        response_removed = True
    else:
        print("[info] No --inventory given: data stays in raw digitizer counts "
              "(no instrument-response removal). Pass --inventory PATH if/when "
              "Sebastien sends a StationXML/dataless for MX.CZB.")

    mseed_path = _process_and_save(st, args, "POPO_CZB", response_removed)
    return mseed_path, st, response_removed


# ---------------------------------------------------------------------------
# Map (single-station location, no country outline asset for Mexico yet)
# ---------------------------------------------------------------------------
# Real, well-known reference points used purely to give the map geographic
# context (city labels + real distances/bearings) -- not an administrative
# boundary survey. Coordinates are public-knowledge city centers.
NEARBY_CITIES = [
    ("Mexico City", 19.4326, -99.1332),
    ("Puebla",      19.0414, -98.2063),
    ("Cholula",     19.0638, -98.3020),
    ("Cuernavaca",  18.9242, -99.2216),
    ("Amecameca",   19.1197, -98.7756),
    ("Atlixco",     18.9096, -98.4347),
]

# Coarser set spanning the whole country, for the "where in Mexico" locator inset.
MEXICO_LOCATOR_CITIES = [
    ("Tijuana",     32.5149, -117.0382),
    ("Monterrey",   25.6866, -100.3161),
    ("Guadalajara", 20.6597, -103.3496),
    ("Merida",      20.9674, -89.5926),
    ("Veracruz",    19.1738, -96.1342),
    ("Oaxaca",      17.0732, -96.7266),
]

COMPASS_DIRS = ["N", "NNE", "NE", "ENE", "E", "ESE", "SE", "SSE",
                "S", "SSW", "SW", "WSW", "W", "WNW", "NW", "NNW"]


def haversine_km(lat1, lon1, lat2, lon2):
    """Great-circle distance in km between two lat/lon points."""
    r_km = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlambda / 2) ** 2
    return 2 * r_km * math.asin(math.sqrt(a))


def bearing_compass(lat1, lon1, lat2, lon2):
    """16-point compass direction from point 1 towards point 2."""
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dlambda = math.radians(lon2 - lon1)
    x = math.sin(dlambda) * math.cos(p2)
    y = math.cos(p1) * math.sin(p2) - math.sin(p1) * math.cos(p2) * math.cos(dlambda)
    brng = (math.degrees(math.atan2(x, y)) + 360) % 360
    return COMPASS_DIRS[int((brng + 11.25) // 22.5) % 16]


def draw_compass_rose(ax, x=0.92, y=0.90, size=0.05):
    ax.annotate(
        "N", xy=(x, y), xytext=(x, y - size), xycoords="axes fraction",
        textcoords="axes fraction", color="white", fontsize=9, weight="bold",
        ha="center", va="center",
        arrowprops=dict(arrowstyle="-|>", color="white", lw=1.4),
    )


def draw_scale_bar(ax, lat_for_scale, km=20.0, x=0.06, y=0.06):
    """Scale bar in data (lon/lat) coordinates, sized so its width equals
    `km` kilometers at the map's latitude."""
    km_per_deg_lon = haversine_km(lat_for_scale, 0.0, lat_for_scale, 1.0)
    width_deg = km / km_per_deg_lon
    x0, x1 = ax.get_xlim()
    y0, y1 = ax.get_ylim()
    bar_x0 = x0 + (x1 - x0) * x
    bar_y = y0 + (y1 - y0) * y
    ax.plot([bar_x0, bar_x0 + width_deg], [bar_y, bar_y], color="white", lw=2.5,
            solid_capstyle="butt", zorder=4)
    ax.text(bar_x0 + width_deg / 2, bar_y, f"{km:.0f} km", color="white", fontsize=7.5,
            ha="center", va="bottom", zorder=4)


def draw_context_map(ax, compact=False):
    """Local-scale map centered on the volcano: real nearby cities with
    computed distance/bearing, compass rose, scale bar. Reused by both the
    standalone `map` action and the combined satellite overview figure."""
    ax.set_facecolor(BG_COLOR)

    for name, lat, lon in NEARBY_CITIES:
        dist_km = haversine_km(VOLCANO_LAT, VOLCANO_LON, lat, lon)
        direction = bearing_compass(VOLCANO_LAT, VOLCANO_LON, lat, lon)
        ax.scatter([lon], [lat], s=40, marker="o", color="#ffd54f",
                   edgecolor="#1c1f26", linewidth=0.6, zorder=3)
        ax.annotate(
            f"{name}\n{dist_km:.0f} km {direction}",
            xy=(lon, lat), xytext=(6, 5), textcoords="offset points",
            color="#ffe082", fontsize=6.5 if compact else 7.5, family="monospace", zorder=3,
        )

    ax.scatter([VOLCANO_LON], [VOLCANO_LAT], s=260, marker="^", color="#ff5252",
               edgecolor="white", linewidth=1.5, zorder=5)
    ax.annotate(
        f"{VOLCANO_NAME}\nsummit ~{VOLCANO_ELEV_M:.0f} m\n(MOUNTS id {MOUNTS_TARGET_ID})",
        xy=(VOLCANO_LON, VOLCANO_LAT), xytext=(12, 10), textcoords="offset points",
        color="white", fontsize=8, family="monospace", zorder=5,
        bbox=dict(boxstyle="round", facecolor="#1c1f26", edgecolor=GRID_COLOR, alpha=0.9),
    )
    if not compact:
        ax.text(
            0.02, 0.02,
            f"Station {NETWORK}.{STATION}: exact coordinates not yet in the data\n"
            f"package (seismic HNZ/HNN/HNE @ loc 00, infrasound HDF @ loc 01-04).\n"
            f"Ask Sebastien Valade for a StationXML/dataless to plot the real\n"
            f"station location and distance to the crater. Nearby city distances/\n"
            f"bearings above are computed (haversine) from public city coordinates.",
            transform=ax.transAxes, color=FG_COLOR, fontsize=7, family="monospace",
            va="bottom", ha="left",
            bbox=dict(boxstyle="round", facecolor="#1c1f26", edgecolor=GRID_COLOR, alpha=0.85),
        )

    pad = 0.3
    ax.set_xlim(VOLCANO_LON - pad, VOLCANO_LON + pad)
    ax.set_ylim(VOLCANO_LAT - pad, VOLCANO_LAT + pad)
    ax.set_aspect(1.0 / max(math.cos(math.radians(VOLCANO_LAT)), 1e-6))
    ax.set_xlabel("Longitude (\u00b0E)", color=FG_COLOR, fontsize=8)
    ax.set_ylabel("Latitude (\u00b0N)", color=FG_COLOR, fontsize=8)
    title = f"{VOLCANO_NAME} -- {NETWORK}.{STATION} context map"
    ax.set_title(title if not compact else title + " (real nearby cities, distances in km)",
                 color="white", fontsize=11 if not compact else 10)
    ax.grid(True, color=GRID_COLOR, linewidth=0.4, alpha=0.5)
    ax.tick_params(colors=FG_COLOR, labelsize=8)
    for spine in ax.spines.values():
        spine.set_color(GRID_COLOR)

    draw_compass_rose(ax)
    scale_x = 0.06 if compact else 0.72
    draw_scale_bar(ax, VOLCANO_LAT, km=20.0, x=scale_x)

    # "Where in Mexico" locator inset -- real coordinates of major cities
    # spanning the country, plotted to true relative scale, with the volcano
    # marked, so the local map above isn't floating in an unlabeled void.
    inset = ax.inset_axes([0.60, 0.60, 0.38, 0.38] if not compact else [0.66, 0.62, 0.32, 0.34])
    inset.set_facecolor("#1c1f26")
    for name, lat, lon in MEXICO_LOCATOR_CITIES:
        inset.scatter([lon], [lat], s=10, color="#80cbc4", zorder=2)
    inset.scatter([VOLCANO_LON], [VOLCANO_LAT], s=70, marker="*", color="#ff5252",
                  edgecolor="white", linewidth=0.5, zorder=3)
    inset.set_xlim(-118, -86)
    inset.set_ylim(14, 33)
    inset.set_aspect(1.0 / max(math.cos(math.radians(23)), 1e-6))
    inset.set_title("Mexico", color=FG_COLOR, fontsize=6.5, pad=2)
    inset.set_xticks([])
    inset.set_yticks([])
    for spine in inset.spines.values():
        spine.set_color(GRID_COLOR)
        spine.set_linewidth(0.5)


def do_map():
    fig, ax = plt.subplots(figsize=(7.5, 7), facecolor=BG_COLOR)
    draw_context_map(ax)
    map_path = os.path.join(MAP_DIR, "station_map.png")
    fig.tight_layout()
    fig.savefig(map_path, dpi=150, facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"[map] Saved station map to {map_path}")
    return map_path


# ---------------------------------------------------------------------------
# Plot (ground data)
# ---------------------------------------------------------------------------
def component_color(channel_code):
    comp = channel_code[-1].upper() if channel_code else "Z"
    return COMPONENT_COLORS.get(comp, "#76ff03")


def rms_envelope(data, sampling_rate, window_s=2.0):
    if len(data) == 0:
        return np.array([], dtype=np.float64)
    window = max(1, min(int(window_s * sampling_rate), len(data)))
    squared = data.astype(np.float64) ** 2
    kernel = np.ones(window) / window
    return np.sqrt(np.convolve(squared, kernel, mode="same"))


def pick_primary_for_spectrogram(st):
    z_traces = [tr for tr in st if tr.stats.channel.upper().endswith("Z")]
    candidates = z_traces or list(st)
    return max(candidates, key=lambda tr: tr.stats.npts / tr.stats.sampling_rate)


def style_axes(ax):
    ax.set_facecolor(BG_COLOR)
    ax.tick_params(colors=FG_COLOR, labelsize=8)
    for spine in ax.spines.values():
        spine.set_color(GRID_COLOR)


def load_fetch_metadata(mseed_path):
    metadata_path = os.path.splitext(mseed_path)[0] + ".json"
    defaults = {"freqmin": 0.5, "freqmax": 10.0, "infra_freqmin": 0.05,
                "infra_freqmax": 20.0, "response_removed": False}
    if not os.path.exists(metadata_path):
        return defaults
    try:
        with open(metadata_path, encoding="utf-8") as f:
            data = json.load(f)
    except (OSError, ValueError):
        return defaults
    defaults.update(data)
    return defaults


def _pick_time_axis_unit(duration_s):
    if duration_s >= 2 * 3600:
        return 3600.0, "Time (hours)", "hours"
    if duration_s >= 180:
        return 60.0, "Time (minutes)", "minutes"
    return 1.0, "Time (seconds)", "seconds"


def do_plot(mseed_path, st=None):
    if st is None:
        st = load_stream(mseed_path)
    base = os.path.splitext(os.path.basename(mseed_path))[0]
    plot_path = os.path.join(PLOT_DIR, f"{base}.png")

    metadata = load_fetch_metadata(mseed_path)
    freqmin, freqmax = metadata["freqmin"], metadata["freqmax"]
    infra_freqmin, infra_freqmax = metadata["infra_freqmin"], metadata["infra_freqmax"]
    response_removed = metadata.get("response_removed", False)

    n_traces = len(st)
    primary = pick_primary_for_spectrogram(st)
    sr = primary.stats.sampling_rate

    t0 = min(tr.stats.starttime for tr in st)
    duration_s = max((tr.stats.starttime - t0) + tr.stats.npts / tr.stats.sampling_rate for tr in st)
    time_divisor, time_label, time_unit_word = _pick_time_axis_unit(duration_s)

    fig = plt.figure(figsize=(15.5, max(6.5, 2.6 + 1.6 * n_traces)), facecolor=BG_COLOR)
    gs = fig.add_gridspec(1 + n_traces, 1, height_ratios=[2.2] + [1] * n_traces, hspace=0.75)

    spec_ax = fig.add_subplot(gs[0])
    nperseg = int(max(32, min(sr * 4, len(primary.data))))
    noverlap = int(nperseg * 0.9)
    f, t, Sxx = spectrogram(primary.data.astype(np.float64), fs=sr, nperseg=nperseg, noverlap=noverlap)
    Sxx_db = 10 * np.log10(Sxx + 1e-30)
    vmin, vmax = np.percentile(Sxx_db, [5, 99.5])
    primary_offset_s = primary.stats.starttime - t0
    mesh = spec_ax.pcolormesh((t + primary_offset_s) / time_divisor, f, Sxx_db, cmap="inferno",
                               shading="auto", vmin=vmin, vmax=vmax)
    is_primary_infra = primary.stats.channel.upper().startswith("HDF")
    primary_freqmax = infra_freqmax if is_primary_infra else freqmax
    y_top = min(sr / 2, max(12.0, primary_freqmax * 1.5))
    spec_ax.set_ylim(0, y_top)
    spec_ax.set_xlim(0, duration_s / time_divisor)
    spec_ax.set_ylabel("Frequency (Hz)", color=FG_COLOR)
    spec_ax.set_title(f"Spectrogram - {primary.id}", color="white", fontsize=10, loc="left")
    spec_ax.set_xticklabels([])
    style_axes(spec_ax)
    cbar = fig.colorbar(mesh, ax=spec_ax, pad=0.01, fraction=0.02)
    cbar.set_label("Power (dB)", color=FG_COLOR)
    cbar.ax.yaxis.set_tick_params(color=FG_COLOR)
    plt.setp(cbar.ax.get_yticklabels(), color=FG_COLOR)

    primary_freqmin = infra_freqmin if is_primary_infra else freqmin
    for edge, label in ((primary_freqmin, "freqmin"), (primary_freqmax, "freqmax")):
        if 0 < edge < y_top:
            spec_ax.axhline(edge, color="white", linewidth=0.8, linestyle="--", alpha=0.6)
            spec_ax.text(0.995, edge, f"{label}={edge:g} Hz",
                         transform=spec_ax.get_yaxis_transform(),
                         color="white", fontsize=7, va="bottom", ha="right",
                         bbox=dict(boxstyle="round,pad=0.15", facecolor=BG_COLOR, edgecolor="none", alpha=0.7))

    mean_power_per_freq = Sxx_db.mean(axis=1)
    peak_freq_hz = f[np.argmax(mean_power_per_freq)]
    if 0 < peak_freq_hz < y_top:
        spec_ax.axhline(peak_freq_hz, color="#39ff14", linewidth=1.0, linestyle="-", alpha=0.8)
        spec_ax.text(0.005, peak_freq_hz, f"peak~{peak_freq_hz:.2g} Hz",
                     transform=spec_ax.get_yaxis_transform(),
                     color="#39ff14", fontsize=7, va="bottom", ha="left",
                     bbox=dict(boxstyle="round,pad=0.15", facecolor=BG_COLOR, edgecolor="none", alpha=0.7))

    max_points = 200_000
    for i, tr in enumerate(st):
        ax = fig.add_subplot(gs[i + 1])
        offset_s = tr.stats.starttime - t0
        trace_duration_s = tr.stats.npts / tr.stats.sampling_rate
        full_data = tr.data.astype(np.float64)
        peak_amp = np.max(np.abs(full_data))
        rms_amp = np.sqrt(np.mean(full_data ** 2))
        times = tr.times() + offset_s
        data = tr.data
        step = max(1, len(data) // max_points)
        if step > 1:
            times = times[::step]
            data = data[::step]

        color = component_color(tr.stats.channel)
        env = rms_envelope(data, tr.stats.sampling_rate / step)
        times_scaled = times / time_divisor
        ax.fill_between(times_scaled, -env, env, color=color, alpha=0.25, linewidth=0)
        ax.plot(times_scaled, data, color=color, linewidth=0.5)
        ax.set_xlim(0, duration_s / time_divisor)

        is_infra = tr.stats.channel.upper().startswith("HDF")
        kind = f"infrasound (loc {tr.stats.location})" if is_infra else "seismic (accelerometer)"
        units = units_for_trace(tr, response_removed)
        line1 = f"{tr.id}  |  {tr.stats.sampling_rate:.0f} Hz  |  {kind}"
        line2_bits = [f"peak {peak_amp:.3e} {units}", f"rms {rms_amp:.3e} {units}"]
        coverage_pct = 100 * trace_duration_s / duration_s
        if coverage_pct < 99:
            end_gap_s = duration_s - (offset_s + trace_duration_s)
            gap_bits = []
            if offset_s > max(1.0, duration_s * 0.01):
                gap_bits.append(f"starts {offset_s:.0f}s late")
            if end_gap_s > max(1.0, duration_s * 0.01):
                gap_bits.append(f"ends {end_gap_s:.0f}s early")
            gap_note = ", ".join(gap_bits) or "partial data"
            line2_bits.append(f"{coverage_pct:.0f}% of window ({gap_note})")
        line2 = "  |  ".join(line2_bits)
        label_text = f"{line1}\n{line2}"
        ax.text(0.0, 1.22, label_text, transform=ax.transAxes, color=color, fontsize=7.8,
                family="monospace", va="bottom", ha="left", linespacing=1.4)
        style_axes(ax)
        if i < n_traces - 1:
            ax.set_xticklabels([])
        else:
            ax.set_xlabel(time_label, color=FG_COLOR)

    start_local_str, tz_label = format_local_time(primary.stats.starttime.datetime)
    end_local_str, _ = format_local_time(primary.stats.endtime.datetime)
    title = (
        f"{VOLCANO_NAME} ({NETWORK}.{STATION})  |  {primary.stats.starttime} - {primary.stats.endtime} UTC  "
        f"({start_local_str} - {end_local_str} {tz_label})  |  seismic bandpass {freqmin:g}-{freqmax:g} Hz, "
        f"infrasound {infra_freqmin:g}-{infra_freqmax:g} Hz"
    )
    fig.suptitle(title, color="white", fontsize=10.5, y=0.995)
    fig.subplots_adjust(left=0.06, right=0.97, top=0.92, bottom=0.06, hspace=0.15)
    fig.savefig(plot_path, dpi=150, facecolor=fig.get_facecolor())
    plt.close(fig)

    print(f"[plot] Saved plot to {plot_path} ({n_traces} traces, "
          f"{duration_s / time_divisor:.1f} {time_unit_word})")
    return plot_path


# ---------------------------------------------------------------------------
# Sonify (ground data) -- "audification": resample the timebase to speed up playback
# ---------------------------------------------------------------------------
def pick_trace(st, channel_filter=None):
    if channel_filter:
        matches = [tr for tr in st if channel_filter.lower() in tr.id.lower()]
        if not matches:
            available = ", ".join(tr.id for tr in st)
            raise SystemExit(f"--channel '{channel_filter}' matched no trace. Available: {available}")
        tr = matches[0]
    else:
        tr = st[0]
    if len(st) > 1:
        others = [t.id for t in st if t.id != tr.id]
        print(f"[sonify] {len(st)} channels available; using '{tr.id}'. "
              f"Ignoring: {', '.join(others)}. Use --channel to pick a different one, "
              f"or --channel all to sonify every one of them.")
    return tr


def _sonify_one_trace(tr, mseed_path, speed_up_factor, channel_tag=None):
    data = tr.data.astype(np.float64)
    data -= data.mean()
    peak = np.max(np.abs(data))
    if peak > 0:
        data /= peak
    audio = (data * 32767).astype(np.int16)

    wav_sample_rate = int(tr.stats.sampling_rate * speed_up_factor)
    input_duration_s = tr.stats.npts / tr.stats.sampling_rate
    output_duration_s = input_duration_s / speed_up_factor

    base = os.path.splitext(os.path.basename(mseed_path))[0]
    if channel_tag:
        wav_path = os.path.join(SONIFY_DIR, f"{base}_{channel_tag}_{int(speed_up_factor)}x.wav")
    else:
        wav_path = os.path.join(SONIFY_DIR, f"{base}_{int(speed_up_factor)}x.wav")
    wavfile.write(wav_path, wav_sample_rate, audio)
    print(f"[sonify] Saved audio to {wav_path}")
    print(f"[sonify]  trace: {tr.id}  |  input: {input_duration_s:.1f}s  ->  "
          f"output: {output_duration_s:.1f}s at {speed_up_factor:.0f}x speed "
          f"(wav sample rate {wav_sample_rate} Hz)")
    return wav_path


def do_sonify(mseed_path, speed_up_factor=DEFAULT_SPEED_UP, channel_filter=None, st=None):
    if st is None:
        st = load_stream(mseed_path)

    if channel_filter and channel_filter.strip().lower() == "all":
        print(f"[sonify] --channel all: sonifying all {len(st)} channel(s) into separate .wav files.")
        wav_paths = []
        seen_tag_counts = {}
        for tr in st:
            channel_tag = tr.id.replace(".", "-")
            seen_tag_counts[channel_tag] = seen_tag_counts.get(channel_tag, 0) + 1
            occurrence = seen_tag_counts[channel_tag]
            if occurrence > 1:
                channel_tag = f"{channel_tag}_seg{occurrence}"
            wav_paths.append(_sonify_one_trace(tr, mseed_path, speed_up_factor, channel_tag))
        return wav_paths

    tr = pick_trace(st, channel_filter)
    return _sonify_one_trace(tr, mseed_path, speed_up_factor)


# ---------------------------------------------------------------------------
# Play
# ---------------------------------------------------------------------------
def do_play(wav_path):
    if not os.path.exists(wav_path):
        print(f"[warn] Cannot play, file not found: {wav_path}", file=sys.stderr)
        return
    print(f"[play] Playing {wav_path}")
    try:
        if sys.platform.startswith("win"):
            import winsound
            winsound.PlaySound(wav_path, winsound.SND_FILENAME)
        elif sys.platform == "darwin":
            subprocess.run(["afplay", wav_path], check=False)
        else:
            for player in ("paplay", "aplay", "ffplay"):
                if shutil.which(player):
                    cmd = [player, wav_path] if player != "ffplay" else [player, "-nodisp", "-autoexit", wav_path]
                    subprocess.run(cmd, check=False)
                    return
            print(f"[play] No audio player found. Open manually: {wav_path}")
    except Exception as exc:
        print(f"[warn] Playback failed for {wav_path}: {exc}", file=sys.stderr)


# ---------------------------------------------------------------------------
# File selection helpers
# ---------------------------------------------------------------------------
def list_mseed_files():
    return sorted(glob.glob(os.path.join(MSEED_DIR, "*.mseed")))


def print_file_list():
    files = list_mseed_files()
    if not files:
        print(f"No .mseed files found in {MSEED_DIR}.")
        return
    print("Saved .mseed files:")
    for i, f in enumerate(files):
        print(f"  [{i}] {os.path.basename(f)}")


def find_wav_for(mseed_path, speed_up_factor, channel_filter=None):
    base = os.path.splitext(os.path.basename(mseed_path))[0]
    if channel_filter and channel_filter.strip().lower() == "all":
        matches = sorted(glob.glob(os.path.join(SONIFY_DIR, f"{base}_*_{int(speed_up_factor)}x.wav")))
        return matches if matches else None
    candidate = os.path.join(SONIFY_DIR, f"{base}_{int(speed_up_factor)}x.wav")
    if os.path.exists(candidate):
        return candidate
    matches = sorted(glob.glob(os.path.join(SONIFY_DIR, f"{base}_[0-9]*x.wav")))
    return matches[-1] if matches else None


def stream_duration_seconds(st):
    tr = st[0]
    return tr.stats.npts / tr.stats.sampling_rate


# ---------------------------------------------------------------------------
# Section B -- MOUNTS satellite API (SO2 / thermal / InSAR)
# ---------------------------------------------------------------------------
MOUNTS_API_URL = "http://mounts-project.com/api/query"
SAT_TYPE_INFO = {
    # type key -> (MOUNTS 'type' query value, display name, color, zero-mapping value)
    "so2":    ("so2",    "SO2 mass (Sentinel-5P)",              "#b388ff", 0.1),
    "mirova": ("mirova", "Thermal hot pixels (Sentinel-2/MIROVA)", "#ff9100", 100000),
    "disp":   ("disp",   "InSAR deformation std.dev (Sentinel-1)", "#ff1744", None),
    "coh":    ("coh",    "InSAR coherence (Sentinel-1)",         "#00b0ff", None),
}


def parse_sat_types_arg(value):
    if value is None:
        value = "all"
    value = value.strip().lower()
    if value == "all":
        return list(SAT_TYPE_INFO)
    tokens = [t.strip() for t in value.split(",") if t.strip()]
    unknown = [t for t in tokens if t not in SAT_TYPE_INFO]
    if unknown:
        raise SystemExit(f"--sat-types: unknown type(s) {unknown}. Known: {sorted(SAT_TYPE_INFO)}")
    return tokens


def mounts_query(target_id, sat_type, time_filter=None, timeout=30):
    """GET the MOUNTS API for one data type. Returns the parsed JSON response
    as-is; normalize_mounts_response() below turns it into (times, values)."""
    params = {
        "table": "results_dat",
        "target_id": str(target_id),
        "type": f"'{sat_type}'",
        "column": "time,data,type",
    }
    if time_filter:
        params["time"] = time_filter
    url = MOUNTS_API_URL + "?" + urllib.parse.urlencode(params, safe="'<>,")
    with urllib.request.urlopen(url, timeout=timeout) as resp:
        raw = resp.read().decode("utf-8")
    try:
        return json.loads(raw)
    except ValueError as exc:
        raise RuntimeError(
            f"MOUNTS API did not return valid JSON for type={sat_type!r} "
            f"(url: {url}): {exc}\nRaw response (truncated): {raw[:300]!r}"
        ) from exc


def normalize_mounts_response(data):
    """Normalize the MOUNTS API JSON into a list of (epoch_seconds, value)
    pairs, sorted by time. Confirmed live (July 2026) against
    http://mounts-project.com/api/query for Popocatepetl (target_id=341090):
    the API returns a JSON list of row dicts, e.g.
    [{"data": 3785.5, "time": "Sun, 25 Feb 2024 20:02:42 GMT", "type": "so2"}, ...].
    The dict-of-parallel-arrays / nested-container branches below are kept as
    defensive fallbacks in case the API's shape changes later -- if parsing
    ever comes up empty, re-run with --sat-debug and inspect the raw saved
    .json under datasets/satellite/raw/."""
    rows = data
    if isinstance(data, dict):
        # Parallel-arrays shape (e.g. {"time": [...], "data": [...], "type": [...]})
        # is checked first and specifically, since the container-key shape below
        # could otherwise ambiguously match on a literal "data" key holding the
        # value array rather than a list of rows.
        if isinstance(data.get("time"), list) and isinstance(data.get("data", data.get("value")), list):
            times = data["time"]
            values = data.get("data", data.get("value"))
            return _finish_normalize(list(zip(times, values)))
        for key in ("results_dat", "rows", "result", "data"):
            if key in data and isinstance(data[key], list) and data[key] and isinstance(data[key][0], (dict, list, tuple)):
                rows = data[key]
                break
        else:
            rows = [data]

    pairs = []
    for row in rows:
        if isinstance(row, dict):
            t = row.get("time")
            v = row.get("data", row.get("value"))
        elif isinstance(row, (list, tuple)):
            t = row[0] if len(row) > 0 else None
            v = row[1] if len(row) > 1 else None
        else:
            continue
        if t is None or v is None:
            continue
        pairs.append((t, v))
    return _finish_normalize(pairs)


def _finish_normalize(pairs):
    out = []
    for t, v in pairs:
        try:
            v = float(v)
        except (TypeError, ValueError):
            continue
        t_epoch = _to_epoch_seconds(t)
        if t_epoch is None:
            continue
        out.append((t_epoch, v))
    out.sort(key=lambda p: p[0])
    return out


def _to_epoch_seconds(t):
    """Parse a MOUNTS API timestamp into epoch seconds. The live API returns
    RFC 2822 dates (e.g. 'Tue, 28 Jul 2026 20:39:34 GMT'), handled here via
    the stdlib email.utils parser; ISO 8601 strings are also accepted as a
    fallback in case a future API revision changes format."""
    if isinstance(t, (int, float)):
        return float(t)
    if isinstance(t, str):
        try:
            dt = email.utils.parsedate_to_datetime(t)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc).timestamp()
        except (TypeError, ValueError):
            pass
        try:
            return UTCDateTime(t).timestamp
        except Exception:
            return None
    return None


def apply_zero_mapping(sat_type, values):
    """Per Sebastien's note: for 'so2', 0.1 maps to 0; for 'mirova', 100000
    maps to 0 (placeholder/no-detection sentinel values used by MOUNTS)."""
    _, _, _, sentinel = SAT_TYPE_INFO[sat_type]
    if sentinel is None:
        return values
    return [0.0 if v == sentinel else v for v in values]


# Maps each --sat-types key to (sheet name, value column header) in a MOUNTS
# Excel export (e.g. MOUNTS_popocatepetl_<daterange>_S1+S2+S5P.xlsx), the full
# historical record Sebastien Valade can export directly from the MOUNTS
# website -- richer than the live API, which only returns recent data.
XLSX_SHEET_MAP = {
    "so2":    ("Sentinel-5P",     "SO2 [tons]"),
    "mirova": ("Sentinel-2",      "SWIR [nb pixels]"),
    "disp":   ("Sentinel-1_disp", "displacement [std.dev m]"),
    "coh":    ("Sentinel-1_coh",  "coherence [nb pixels]"),
}


def find_default_sat_xlsx():
    """Auto-detect a MOUNTS_*.xlsx export dropped into datasets/ (not tracked
    in git -- see .gitignore). Preferred over the live API when present, since
    it's the complete historical record rather than just recent data."""
    candidates = sorted(glob.glob(os.path.join(BASE_DIR, "datasets", "MOUNTS_*.xlsx")))
    return candidates[0] if candidates else None


def load_mounts_xlsx(path, sat_types, sat_start=None, sat_end=None):
    """Load one or more series directly from a MOUNTS Excel export instead of
    the live API. Expects one sheet per data type, per XLSX_SHEET_MAP."""
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit(
            "Reading a MOUNTS .xlsx export requires the 'openpyxl' package: "
            "pip install openpyxl"
        ) from exc

    start_dt = datetime.strptime(sat_start, "%Y-%m-%d") if sat_start else None
    end_dt = datetime.strptime(sat_end, "%Y-%m-%d") if sat_end else None

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    series = {}
    for sat_type in sat_types:
        sheet_name, value_col = XLSX_SHEET_MAP[sat_type]
        if sheet_name not in wb.sheetnames:
            print(f"[satellite] Sheet {sheet_name!r} not found in {path}; skipping {sat_type}.",
                  file=sys.stderr)
            continue
        rows = list(wb[sheet_name].iter_rows(values_only=True))
        header = rows[0]
        if "date" not in header or value_col not in header:
            print(f"[satellite] Expected columns 'date'/{value_col!r} not found in sheet "
                  f"{sheet_name!r}; skipping {sat_type}.", file=sys.stderr)
            continue
        date_idx, value_idx = header.index("date"), header.index(value_col)

        pairs = []
        for row in rows[1:]:
            dt, v = row[date_idx], row[value_idx]
            if dt is None or v is None:
                continue
            if start_dt and dt < start_dt:
                continue
            if end_dt and dt > end_dt:
                continue
            pairs.append((dt.replace(tzinfo=timezone.utc).timestamp(), float(v)))
        pairs.sort(key=lambda p: p[0])
        if not pairs:
            print(f"[satellite] No rows for {sat_type} in the requested date range; skipping.",
                  file=sys.stderr)
            continue
        times, values = zip(*pairs)
        values = apply_zero_mapping(sat_type, list(values))
        series[sat_type] = (list(times), values)
        print(f"[satellite] {sat_type}: {len(times)} point(s) from {os.path.basename(path)}, "
              f"{datetime.utcfromtimestamp(times[0])} - {datetime.utcfromtimestamp(times[-1])} UTC")
    if not series:
        raise SystemExit(f"No usable series loaded from {path}.")
    return series


def do_satellite_fetch(sat_types, target_id, time_filter, debug=False):
    series = {}
    for sat_type in sat_types:
        api_type, name, color, _sentinel = SAT_TYPE_INFO[sat_type]
        print(f"[satellite] Querying MOUNTS for {name} (type={api_type!r}, target_id={target_id})...")
        try:
            raw = mounts_query(target_id, api_type, time_filter)
        except Exception as exc:
            print(f"[satellite] Failed to fetch {sat_type}: {exc}", file=sys.stderr)
            continue

        raw_path = os.path.join(SAT_RAW_DIR, f"popo_{sat_type}.json")
        with open(raw_path, "w", encoding="utf-8") as f:
            json.dump(raw, f, indent=2)
        if debug:
            print(f"[satellite] Raw response saved to {raw_path}")

        pairs = normalize_mounts_response(raw)
        if not pairs:
            print(f"[satellite] No usable (time, data) pairs parsed for {sat_type} "
                  f"-- see {raw_path} and adjust normalize_mounts_response() if the "
                  f"API's JSON shape differs from what was assumed.", file=sys.stderr)
            continue
        times, values = zip(*pairs)
        values = apply_zero_mapping(sat_type, values)
        series[sat_type] = (list(times), list(values))
        print(f"[satellite] {sat_type}: {len(times)} point(s), "
              f"{datetime.utcfromtimestamp(times[0])} - {datetime.utcfromtimestamp(times[-1])} UTC")
    if not series:
        raise SystemExit("No satellite series could be fetched/parsed. Check network access "
                          "and see the [satellite] warnings above.")
    return series


def do_satellite_plot(series, target_id):
    """Combined overview figure: the MOUNTS time series stacked on the left,
    a real-geography context map on the right, so location and activity read
    as one explanation instead of two disconnected images."""
    n = len(series)
    fig = plt.figure(figsize=(14.5, max(7.0, 1.9 * n + 1.2)), facecolor=BG_COLOR)
    gs = fig.add_gridspec(n, 2, width_ratios=[2.1, 1.5], wspace=0.25, hspace=0.55)

    for i, sat_type in enumerate(series):
        ax = fig.add_subplot(gs[i, 0])
        times, values = series[sat_type]
        _, name, color, _ = SAT_TYPE_INFO[sat_type]
        dt = [datetime.utcfromtimestamp(t) for t in times]
        ax.plot(dt, values, color=color, marker="o", markersize=3, linewidth=1.0)
        ax.set_title(name, color="white", fontsize=10, loc="left")
        style_axes(ax)
        ax.grid(True, color=GRID_COLOR, linewidth=0.4, alpha=0.5)

    map_ax = fig.add_subplot(gs[:, 1])
    draw_context_map(map_ax, compact=True)

    fig.suptitle(f"{VOLCANO_NAME} -- MOUNTS satellite time series + location (target_id {target_id})",
                 color="white", fontsize=12)
    # Not tight_layout(): inset_axes (the Mexico locator) isn't compatible with it
    # and produces a large blank gap -- position panels manually instead.
    fig.subplots_adjust(top=0.93, bottom=0.06, left=0.06, right=0.98, wspace=0.25, hspace=0.55)
    plot_path = os.path.join(SAT_PLOT_DIR, "popo_satellite.png")
    fig.savefig(plot_path, dpi=150, facecolor=fig.get_facecolor())
    plt.close(fig)
    print(f"[satellite] Saved plot to {plot_path}")
    return plot_path


# Base carrier-frequency range per data type, chosen so a mixed-down piece
# keeps the four series distinguishable by ear (deformation low/ominous,
# coherence high/shimmering, SO2 and thermal in between).
SAT_FREQ_RANGE_HZ = {
    "disp": (80, 220),
    "so2": (220, 440),
    "mirova": (440, 880),
    "coh": (880, 1760),
}


def sonify_timeseries(sat_type, times, values, listen_minutes, sample_rate=44100, grain_s=0.15):
    """Parameter-mapping sonification: each (time, value) point becomes a
    short pitched grain. Pitch <- value (log-mapped within this series'
    observed range); position in the output audio <- measurement time,
    linearly compressed so the whole series fits in --listen-minutes."""
    freq_lo, freq_hi = SAT_FREQ_RANGE_HZ[sat_type]
    t_arr = np.asarray(times, dtype=np.float64)
    v_arr = np.asarray(values, dtype=np.float64)

    t_min, t_max = t_arr.min(), t_arr.max()
    t_span = max(t_max - t_min, 1.0)
    v_min, v_max = v_arr.min(), v_arr.max()
    v_span = v_max - v_min

    total_s = listen_minutes * 60.0
    grain_n = max(1, int(grain_s * sample_rate))
    n_samples = int(total_s * sample_rate) + grain_n + 1
    buffer = np.zeros(n_samples, dtype=np.float64)

    # Grain time base + envelope are identical for every point (only the
    # carrier frequency changes) -- compute them once instead of rebuilding
    # an arange()/sin() envelope on every one of the (up to several thousand) points.
    t_grain = np.arange(grain_n) / sample_rate
    envelope = 0.6 * np.sin(np.pi * t_grain / grain_s) ** 2  # smooth in/out, no clicks

    # Vectorize the per-point pitch/position mapping across the whole series
    # instead of doing it one Python scalar at a time inside the loop.
    norm_v = np.full_like(v_arr, 0.5) if v_span == 0 else np.clip((v_arr - v_min) / v_span, 0.0, 1.0)
    freqs = freq_lo * (freq_hi / freq_lo) ** norm_v
    start_samples = ((t_arr - t_min) / t_span * total_s * sample_rate).astype(np.int64)

    for freq, start_sample in zip(freqs, start_samples):
        grain = envelope * np.sin(2 * np.pi * freq * t_grain)
        end_sample = start_sample + grain_n
        if end_sample > len(buffer):
            grain = grain[: len(buffer) - start_sample]
            end_sample = len(buffer)
        buffer[start_sample:end_sample] += grain

    peak = np.max(np.abs(buffer))
    if peak > 0:
        buffer /= peak
    audio = (buffer * 32767 * 0.9).astype(np.int16)

    wav_path = os.path.join(SAT_SONIFY_DIR, f"popo_{sat_type}.wav")
    wavfile.write(wav_path, sample_rate, audio)
    print(f"[satellite] Saved sonification to {wav_path} "
          f"({len(t_arr)} points -> {total_s:.0f}s of audio, {freq_lo:.0f}-{freq_hi:.0f} Hz range)")
    return wav_path, audio, sample_rate


def export_envelope_csv(sat_type, times, values, listen_minutes, fps=30):
    """Write a uniformly-resampled 0..1 control curve (CSV: time_s, value_norm,
    value_raw) for this series, on the exact same compressed timeline used by
    sonify_timeseries() -- so a REAPER volume/FX envelope or a GPIO LED
    brightness script driven by this file stays in sync with the .wav.

    value_norm is linearly interpolated between the real (sparse, irregular)
    measurements, so it's a smooth continuous curve rather than the discrete
    pitched grains used for audio -- suitable for driving a fader or a PWM
    LED at a fixed frame rate (default 30 fps)."""
    t_arr = np.asarray(times, dtype=np.float64)
    v_arr = np.asarray(values, dtype=np.float64)

    order = np.argsort(t_arr)
    t_arr = t_arr[order]
    v_arr = v_arr[order]

    t_min, t_max = t_arr.min(), t_arr.max()
    t_span = max(t_max - t_min, 1.0)
    v_min, v_max = v_arr.min(), v_arr.max()
    v_span = v_max - v_min

    total_s = listen_minutes * 60.0
    pos_s = (t_arr - t_min) / t_span * total_s
    norm_v = np.full_like(v_arr, 0.5) if v_span == 0 else np.clip((v_arr - v_min) / v_span, 0.0, 1.0)

    n_frames = max(2, int(total_s * fps) + 1)
    grid_s = np.linspace(0.0, total_s, n_frames)
    # np.interp holds the nearest endpoint's value outside [pos_s[0], pos_s[-1]],
    # so the curve never extrapolates wildly before the first / after the last point.
    grid_norm = np.interp(grid_s, pos_s, norm_v)
    grid_raw = np.interp(grid_s, pos_s, v_arr)

    csv_path = os.path.join(SAT_ENVELOPE_DIR, f"popo_{sat_type}_envelope.csv")
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["time_s", "value_norm", "value_raw"])
        for s, n, r in zip(grid_s, grid_norm, grid_raw):
            writer.writerow([f"{s:.3f}", f"{n:.4f}", f"{r:.6g}"])
    print(f"[satellite] Saved control-curve envelope to {csv_path} "
          f"({n_frames} points @ {fps} fps, {total_s:.0f}s)")
    return csv_path


def do_satellite_sonify(series, listen_minutes, sample_rate=44100):
    wav_paths = {}
    csv_paths = {}
    mix = None
    for sat_type, (times, values) in series.items():
        wav_path, audio, sr = sonify_timeseries(sat_type, times, values, listen_minutes, sample_rate)
        wav_paths[sat_type] = wav_path
        csv_paths[sat_type] = export_envelope_csv(sat_type, times, values, listen_minutes)
        audio_f = audio.astype(np.float64)
        if mix is None:
            mix = audio_f.copy()
        else:
            if len(audio_f) > len(mix):
                mix = np.pad(mix, (0, len(audio_f) - len(mix)))
            elif len(mix) > len(audio_f):
                audio_f = np.pad(audio_f, (0, len(mix) - len(audio_f)))
            mix += audio_f
    if mix is not None:
        peak = np.max(np.abs(mix))
        if peak > 0:
            mix = mix / peak * 32767 * 0.9
        mix_path = os.path.join(SAT_SONIFY_DIR, "popo_satellite_mix.wav")
        wavfile.write(mix_path, sample_rate, mix.astype(np.int16))
        print(f"[satellite] Saved combined mix to {mix_path}")
        wav_paths["mix"] = mix_path
    return wav_paths, csv_paths


def do_satellite(args):
    sat_types = parse_sat_types_arg(args.sat_types)

    xlsx_path = args.sat_xlsx or (None if args.sat_live else find_default_sat_xlsx())
    if xlsx_path:
        print(f"[satellite] Using local MOUNTS export: {xlsx_path} "
              f"(pass --sat-live to force the live API instead)")
        series = load_mounts_xlsx(xlsx_path, sat_types, args.sat_start, args.sat_end)
    else:
        time_filter = None
        if args.sat_start or args.sat_end:
            parts = []
            if args.sat_start:
                parts.append(f">{args.sat_start}")
            if args.sat_end:
                parts.append(f"<{args.sat_end}")
            time_filter = ",".join(parts)
        elif args.sat_days_back:
            start = (datetime.utcnow() - timedelta(days=args.sat_days_back)).strftime("%Y-%m-%d")
            time_filter = f">{start}"
        series = do_satellite_fetch(sat_types, args.target_id, time_filter, debug=args.sat_debug)

    do_satellite_plot(series, args.target_id)
    wav_paths, _csv_paths = do_satellite_sonify(series, args.listen_minutes or 2.0)
    if args.play:
        for p in wav_paths.values():
            do_play(p)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
VALID_ACTIONS = {"fetch", "plot", "sonify", "play", "map", "satellite", "all"}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Fetch, plot and/or sonify Popocatepetl (MX.CZB) seismic/infrasound data, "
                    "and/or MOUNTS satellite time series.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "actions", nargs="*", default=None,
        help="What to do: any combination of fetch, plot, sonify, play, map, satellite, all "
             "(default: fetch plot sonify)",
    )

    ground_group = parser.add_argument_group("ground data (SDS archive)")
    ground_group.add_argument("--sds-root", default=DEFAULT_SDS_ROOT, metavar="PATH",
                               help=f"Local SDS archive root (default: {DEFAULT_SDS_ROOT}). "
                                    "Unzip Sebastien's WeTransfer package here.")
    ground_group.add_argument("--channels", default="all", metavar="SEISMIC|INFRASOUND|ALL|LIST",
                               help="Which channels to fetch: 'seismic' (HNZ/HNN/HNE), "
                                    "'infrasound' (HDF01-04), 'all' (default), or a comma-list "
                                    f"of tokens from {sorted(CHANNEL_TOKEN_MAP)}.")
    ground_group.add_argument("--inventory", default=None, metavar="PATH",
                               help="StationXML/dataless for MX.CZB, if available, to remove the "
                                    "instrument response. Without it, data stays in raw counts.")

    window_group = parser.add_argument_group("time window (data only covers 2026-03-27)")
    window_group.add_argument("--date", default=DEFAULT_DATE, help=f"Date, YYYY-MM-DD (default: {DEFAULT_DATE})")
    window_group.add_argument("--start", default=None, metavar="HH:MM", help="Window start, UTC")
    window_group.add_argument("--end", default=None, metavar="HH:MM", help="Window end, UTC")
    window_group.add_argument("--full-day", action="store_true",
                               help="Use the entire archived day instead of a window")
    window_group.add_argument("--before-min", type=float, default=DEFAULT_BEFORE_MIN,
                               help=f"Minutes before the {DEFAULT_ERUPTION_UTC_HHMM} UTC eruption to start "
                                    f"the default window (default: {DEFAULT_BEFORE_MIN:.0f})")
    window_group.add_argument("--after-min", type=float, default=DEFAULT_AFTER_MIN,
                               help=f"Minutes after the eruption to end the default window "
                                    f"(default: {DEFAULT_AFTER_MIN:.0f})")

    filter_group = parser.add_argument_group("bandpass filters")
    filter_group.add_argument("--freqmin", type=float, default=0.5, help="Seismic lower corner, Hz (default: 0.5)")
    filter_group.add_argument("--freqmax", type=float, default=10.0, help="Seismic upper corner, Hz (default: 10.0)")
    filter_group.add_argument("--infra-freqmin", type=float, default=0.05,
                               help="Infrasound lower corner, Hz (default: 0.05)")
    filter_group.add_argument("--infra-freqmax", type=float, default=20.0,
                               help="Infrasound upper corner, Hz (default: 20.0)")

    sonify_group = parser.add_argument_group("sonify options")
    sonify_group.add_argument("--speed-up", type=float, default=None,
                               help=f"Playback speed multiplier (default: {DEFAULT_SPEED_UP})")
    sonify_group.add_argument("--channel", default=None, metavar="SEED_ID_SUBSTRING",
                               help="Which trace to sonify (substring match, e.g. 'HNZ' or 'HDF.01'). "
                                    "Pass 'all' to sonify every channel into its own .wav.")
    sonify_group.add_argument("--listen-minutes", type=float, default=None, metavar="MINUTES",
                               help="Ground: target sonification length in minutes. "
                                    "Satellite: length of the parameter-mapped piece (default: 2).")

    select_group = parser.add_argument_group("file selection (skip fetching)")
    select_group.add_argument("--list", action="store_true", help="List saved .mseed files and exit")
    select_group.add_argument("--pick", type=int, default=None, metavar="INDEX",
                               help="Use the .mseed file at INDEX from --list instead of fetching")
    select_group.add_argument("--file", default=None, help="Use a specific .mseed file path instead of fetching")

    sat_group = parser.add_argument_group("satellite data (MOUNTS API)")
    sat_group.add_argument("--target-id", type=int, default=MOUNTS_TARGET_ID,
                            help=f"MOUNTS/Smithsonian volcano id (default: {MOUNTS_TARGET_ID}, Popocatepetl)")
    sat_group.add_argument("--sat-types", default="all", metavar="LIST",
                            help=f"Comma-list from {sorted(SAT_TYPE_INFO)}, or 'all' (default)")
    sat_group.add_argument("--sat-start", default=None, metavar="YYYY-MM-DD", help="Satellite series start date")
    sat_group.add_argument("--sat-end", default=None, metavar="YYYY-MM-DD", help="Satellite series end date")
    sat_group.add_argument("--sat-days-back", type=int, default=365,
                            help="If --sat-start/--sat-end aren't given, fetch this many days back "
                                 "from now (default: 365)")
    sat_group.add_argument("--sat-debug", action="store_true",
                            help="Print/keep the raw MOUNTS JSON response for inspection")
    sat_group.add_argument("--sat-xlsx", default=None, metavar="PATH",
                            help="Load series from a local MOUNTS Excel export instead of the "
                                 "live API (auto-detected if a datasets/MOUNTS_*.xlsx file exists).")
    sat_group.add_argument("--sat-live", action="store_true",
                            help="Force fetching from the live MOUNTS API even if a local "
                                 "datasets/MOUNTS_*.xlsx export is auto-detected.")
    sat_group.add_argument("--play", action="store_true", help="Play the satellite .wav file(s) after sonifying")

    return parser.parse_args()


def main():
    args = parse_args()

    if args.actions:
        invalid = set(args.actions) - VALID_ACTIONS
        if invalid:
            raise SystemExit(f"Invalid action(s): {', '.join(sorted(invalid))}. "
                              f"Choose from: {', '.join(sorted(VALID_ACTIONS))}")

    if args.freqmin <= 0 or args.infra_freqmin <= 0:
        raise SystemExit("--freqmin/--infra-freqmin must be > 0")
    if args.freqmax <= args.freqmin:
        raise SystemExit(f"--freqmax ({args.freqmax}) must be > --freqmin ({args.freqmin})")
    if args.infra_freqmax <= args.infra_freqmin:
        raise SystemExit(f"--infra-freqmax ({args.infra_freqmax}) must be > --infra-freqmin ({args.infra_freqmin})")
    if args.speed_up is not None and args.speed_up <= 0:
        raise SystemExit(f"--speed-up must be > 0 (got {args.speed_up})")
    if args.listen_minutes is not None and args.listen_minutes <= 0:
        raise SystemExit(f"--listen-minutes must be > 0 (got {args.listen_minutes})")

    if args.list:
        print_file_list()
        return

    actions = set(args.actions) if args.actions else {"fetch", "plot", "sonify"}
    if "all" in actions:
        actions = {"fetch", "plot", "sonify"}

    if "map" in actions:
        do_map()
        actions.discard("map")
        if not actions:
            return

    if "satellite" in actions:
        do_satellite(args)
        actions.discard("satellite")
        if not actions:
            return

    mseed_path = None
    response_removed = False
    if args.file:
        if not os.path.exists(args.file):
            raise SystemExit(f"--file not found: {args.file}")
        mseed_path = args.file
        actions.discard("fetch")
    elif args.pick is not None:
        files = list_mseed_files()
        if not files or not (0 <= args.pick < len(files)):
            print_file_list()
            raise SystemExit(f"Invalid --pick index: {args.pick}")
        mseed_path = files[args.pick]
        actions.discard("fetch")

    fetch_needed = "fetch" in actions
    speed_up_explicit = args.speed_up is not None

    fetched_st = None
    if fetch_needed:
        mseed_path, fetched_st, response_removed = do_fetch(args)
    elif mseed_path is None:
        files = list_mseed_files()
        if not files:
            raise SystemExit("No .mseed files available and 'fetch' not requested. "
                              "Run with 'fetch' or use --file/--pick.")
        mseed_path = files[-1]
        print(f"[info] Using latest saved file: {mseed_path}")

    preloaded_st = None
    if ("plot" in actions or "sonify" in actions or "play" in actions) and not fetch_needed:
        preloaded_st = load_stream(mseed_path)
        describe_stream(preloaded_st, prefix="info")
    current_st = fetched_st if fetched_st is not None else preloaded_st

    if args.listen_minutes and not fetch_needed:
        st_for_duration = current_st if current_st is not None else load_stream(mseed_path)
        input_duration_s = stream_duration_seconds(st_for_duration)
        target_s = args.listen_minutes * 60
        args.speed_up = input_duration_s / target_s
        print(f"[info] Existing recording is {input_duration_s / 60:.1f} min long; using "
              f"speed-up {args.speed_up:.2f}x to produce a {args.listen_minutes:.1f} min sonification.")
        if args.speed_up < MIN_LISTEN_SPEED_UP:
            print(f"[warn] {args.speed_up:.2f}x is below the ~{MIN_LISTEN_SPEED_UP:.0f}x floor for "
                  f"audible ground motion; this piece will likely sound like a faint rumble.",
                  file=sys.stderr)
    elif args.listen_minutes and fetch_needed:
        args.speed_up = args.speed_up if speed_up_explicit else MIN_LISTEN_SPEED_UP
    elif args.speed_up is None:
        args.speed_up = DEFAULT_SPEED_UP

    if "plot" in actions:
        do_plot(mseed_path, st=current_st)

    wav_path = None
    if "sonify" in actions:
        wav_path = do_sonify(mseed_path, args.speed_up, args.channel, st=current_st)

    if "play" in actions:
        if wav_path is None:
            wav_path = find_wav_for(mseed_path, args.speed_up, args.channel)
        if wav_path is None:
            wav_path = do_sonify(mseed_path, args.speed_up, args.channel, st=current_st)
        for p in (wav_path if isinstance(wav_path, list) else [wav_path]):
            do_play(p)


if __name__ == "__main__":
    main()
